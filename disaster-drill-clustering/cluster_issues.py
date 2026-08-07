#!/usr/bin/env python3
"""Explore Japanese disaster-drill issues with sentence embeddings and clustering."""

from __future__ import annotations

import argparse
import csv
import json
import math
import sys
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from shutil import copy2
from typing import Sequence

import numpy as np
from openpyxl import Workbook, load_workbook
from sklearn.cluster import AgglomerativeClustering
from sklearn.metrics import silhouette_score
from sklearn.preprocessing import normalize


DEFAULT_MODEL = "sentence-transformers/paraphrase-multilingual-mpnet-base-v2"
DEFAULT_CLUSTER_COUNTS = (8, 10, 12, 15, 20)


@dataclass(frozen=True)
class IssueRow:
    excel_row: int
    text: str


@dataclass(frozen=True)
class ClusterEvaluation:
    n_clusters: int
    silhouette_cosine: float
    smallest_cluster: int
    largest_cluster: int
    cluster_size_std: float


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="日本語の課題文を埋め込み、階層クラスタリングを比較します。"
    )
    parser.add_argument("input", type=Path, help="入力Excelファイル")
    parser.add_argument("--output-dir", type=Path, default=Path("outputs"))
    parser.add_argument("--sheet", help="対象シート名。省略時は最初のシート")
    parser.add_argument("--text-column", default="課題", help="課題文の列名")
    parser.add_argument(
        "--cluster-counts",
        nargs="+",
        type=int,
        default=list(DEFAULT_CLUSTER_COUNTS),
        metavar="K",
    )
    parser.add_argument(
        "--selected-k",
        default="auto",
        help="結果Excelに採用するクラスタ数。autoまたは候補内の整数",
    )
    parser.add_argument("--model", default=DEFAULT_MODEL)
    parser.add_argument("--batch-size", type=int, default=32)
    parser.add_argument("--device", help="例: cpu, cuda, mps。省略時は自動判定")
    parser.add_argument("--representatives", type=int, default=5)
    parser.add_argument("--random-state", type=int, default=42)
    parser.add_argument("--umap-neighbors", type=int, default=15)
    parser.add_argument(
        "--force-embedding",
        action="store_true",
        help="キャッシュ済み埋め込みを使わず再計算する",
    )
    return parser.parse_args()


def read_issues(
    input_path: Path, sheet_name: str | None, text_column: str
) -> tuple[str, list[IssueRow]]:
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    try:
        if sheet_name is None:
            worksheet = workbook[workbook.sheetnames[0]]
        elif sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
        else:
            raise ValueError(
                f"シート {sheet_name!r} がありません。候補: {workbook.sheetnames}"
            )

        headers = {
            str(cell.value).strip(): cell.column
            for cell in worksheet[1]
            if cell.value is not None
        }
        if text_column not in headers:
            raise ValueError(
                f"列 {text_column!r} がありません。候補: {list(headers)}"
            )

        column_index = headers[text_column]
        rows: list[IssueRow] = []
        for row_number in range(2, worksheet.max_row + 1):
            value = worksheet.cell(row=row_number, column=column_index).value
            if value is None:
                continue
            text = str(value).strip()
            if text:
                rows.append(IssueRow(excel_row=row_number, text=text))
        return worksheet.title, rows
    finally:
        workbook.close()


def validate_cluster_counts(cluster_counts: Sequence[int], n_samples: int) -> list[int]:
    counts = sorted(set(cluster_counts))
    if not counts:
        raise ValueError("クラスタ数を1つ以上指定してください。")
    invalid = [count for count in counts if count < 2 or count >= n_samples]
    if invalid:
        raise ValueError(
            f"クラスタ数は2以上、課題数({n_samples})未満にしてください: {invalid}"
        )
    return counts


def embedding_cache_key(texts: Sequence[str], model_name: str) -> str:
    import hashlib

    digest = hashlib.sha256()
    digest.update(model_name.encode("utf-8"))
    for text in texts:
        digest.update(b"\0")
        digest.update(text.encode("utf-8"))
    return digest.hexdigest()


def load_or_create_embeddings(
    texts: Sequence[str],
    model_name: str,
    output_dir: Path,
    batch_size: int,
    device: str | None,
    force: bool,
) -> np.ndarray:
    from sentence_transformers import SentenceTransformer

    cache_path = output_dir / "embeddings.npz"
    expected_key = embedding_cache_key(texts, model_name)
    if cache_path.exists() and not force:
        with np.load(cache_path, allow_pickle=False) as cached:
            cached_key = str(cached["cache_key"].item())
            if cached_key == expected_key:
                print(f"埋め込みキャッシュを使用: {cache_path}")
                return np.asarray(cached["embeddings"], dtype=np.float32)

    print(f"モデルを読み込み中: {model_name}")
    model = SentenceTransformer(model_name, device=device)
    embeddings = model.encode(
        list(texts),
        batch_size=batch_size,
        show_progress_bar=True,
        convert_to_numpy=True,
        normalize_embeddings=True,
    )
    embeddings = np.asarray(embeddings, dtype=np.float32)
    np.savez_compressed(
        cache_path, embeddings=embeddings, cache_key=np.array(expected_key)
    )
    return embeddings


def fit_clusters(embeddings: np.ndarray, n_clusters: int) -> np.ndarray:
    # 正規化済みベクトルではcosine距離が安定して扱える。
    normalized = normalize(embeddings, norm="l2")
    model = AgglomerativeClustering(
        n_clusters=n_clusters,
        metric="cosine",
        linkage="average",
    )
    return model.fit_predict(normalized)


def evaluate_clustering(
    embeddings: np.ndarray, labels: np.ndarray, n_clusters: int
) -> ClusterEvaluation:
    sizes = np.bincount(labels, minlength=n_clusters)
    return ClusterEvaluation(
        n_clusters=n_clusters,
        silhouette_cosine=float(
            silhouette_score(embeddings, labels, metric="cosine")
        ),
        smallest_cluster=int(sizes.min()),
        largest_cluster=int(sizes.max()),
        cluster_size_std=float(sizes.std()),
    )


def representative_indices(
    embeddings: np.ndarray,
    labels: np.ndarray,
    n_representatives: int,
) -> dict[int, list[tuple[int, float]]]:
    normalized = normalize(embeddings, norm="l2")
    result: dict[int, list[tuple[int, float]]] = {}
    for cluster_id in sorted(int(value) for value in np.unique(labels)):
        member_indices = np.flatnonzero(labels == cluster_id)
        centroid = normalized[member_indices].mean(axis=0)
        centroid_norm = np.linalg.norm(centroid)
        if centroid_norm:
            centroid = centroid / centroid_norm
        similarities = normalized[member_indices] @ centroid
        order = np.argsort(-similarities, kind="stable")[:n_representatives]
        result[cluster_id] = [
            (int(member_indices[position]), float(similarities[position]))
            for position in order
        ]
    return result


def create_umap(
    embeddings: np.ndarray, random_state: int, n_neighbors: int
) -> np.ndarray:
    import umap

    safe_neighbors = min(max(2, n_neighbors), len(embeddings) - 1)
    reducer = umap.UMAP(
        n_components=2,
        n_neighbors=safe_neighbors,
        metric="cosine",
        min_dist=0.1,
        random_state=random_state,
    )
    return np.asarray(reducer.fit_transform(embeddings))


def write_csv(path: Path, fieldnames: Sequence[str], rows: Sequence[dict]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_evaluation_csv(
    path: Path, evaluations: Sequence[ClusterEvaluation]
) -> None:
    write_csv(
        path,
        list(asdict(evaluations[0]).keys()),
        [asdict(evaluation) for evaluation in evaluations],
    )


def write_assignments_csv(
    path: Path,
    issues: Sequence[IssueRow],
    all_labels: dict[int, np.ndarray],
    coordinates: np.ndarray,
) -> None:
    fieldnames = ["excel_row", "課題"] + [
        f"cluster_{count}" for count in sorted(all_labels)
    ] + ["umap_x", "umap_y"]
    rows = []
    for index, issue in enumerate(issues):
        row = {"excel_row": issue.excel_row, "課題": issue.text}
        row.update(
            {
                f"cluster_{count}": int(labels[index])
                for count, labels in sorted(all_labels.items())
            }
        )
        row.update(umap_x=float(coordinates[index, 0]), umap_y=float(coordinates[index, 1]))
        rows.append(row)
    write_csv(path, fieldnames, rows)


def write_cluster_summary(
    path: Path,
    issues: Sequence[IssueRow],
    embeddings: np.ndarray,
    labels: np.ndarray,
    selected_k: int,
    n_representatives: int,
) -> None:
    representatives = representative_indices(embeddings, labels, n_representatives)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "cluster_summary"
    worksheet.append(
        [
            "cluster",
            "cluster_name",
            "size",
            "representative_rank",
            "centroid_similarity",
            "excel_row",
            "representative_issue",
        ]
    )
    for cluster_id in range(selected_k):
        size = int(np.count_nonzero(labels == cluster_id))
        for rank, (issue_index, similarity) in enumerate(
            representatives[cluster_id], start=1
        ):
            issue = issues[issue_index]
            worksheet.append(
                [
                    cluster_id,
                    "",
                    size,
                    rank,
                    similarity,
                    issue.excel_row,
                    issue.text,
                ]
            )
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.column_dimensions["A"].width = 12
    worksheet.column_dimensions["B"].width = 28
    worksheet.column_dimensions["C"].width = 10
    worksheet.column_dimensions["D"].width = 22
    worksheet.column_dimensions["E"].width = 22
    worksheet.column_dimensions["F"].width = 12
    worksheet.column_dimensions["G"].width = 90
    workbook.save(path)


def append_clusters_to_excel(
    input_path: Path,
    output_path: Path,
    sheet_name: str,
    issues: Sequence[IssueRow],
    labels: np.ndarray,
) -> None:
    copy2(input_path, output_path)
    workbook = load_workbook(output_path)
    worksheet = workbook[sheet_name]
    cluster_column = worksheet.max_column + 1
    name_column = cluster_column + 1
    worksheet.cell(row=1, column=cluster_column, value="cluster")
    worksheet.cell(row=1, column=name_column, value="cluster_name")
    for issue, label in zip(issues, labels, strict=True):
        worksheet.cell(row=issue.excel_row, column=cluster_column, value=int(label))
        worksheet.cell(row=issue.excel_row, column=name_column, value="")
    workbook.save(output_path)
    workbook.close()


def plot_comparison(
    path: Path,
    coordinates: np.ndarray,
    all_labels: dict[int, np.ndarray],
    evaluations: Sequence[ClusterEvaluation],
) -> None:
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    counts = sorted(all_labels)
    columns = min(3, len(counts))
    rows = math.ceil(len(counts) / columns)
    figure, axes = plt.subplots(
        rows, columns, figsize=(6 * columns, 5 * rows), squeeze=False
    )
    score_by_k = {item.n_clusters: item.silhouette_cosine for item in evaluations}
    for axis, count in zip(axes.flat, counts):
        labels = all_labels[count]
        scatter = axis.scatter(
            coordinates[:, 0],
            coordinates[:, 1],
            c=labels,
            cmap="tab20",
            s=18,
            alpha=0.8,
        )
        axis.set_title(f"k={count}  silhouette={score_by_k[count]:.3f}")
        axis.set_xlabel("UMAP 1")
        axis.set_ylabel("UMAP 2")
        figure.colorbar(scatter, ax=axis, shrink=0.75, label="cluster")
    for axis in axes.flat[len(counts) :]:
        axis.set_visible(False)
    figure.suptitle("Disaster-drill issues: clustering comparison", fontsize=15)
    figure.tight_layout()
    figure.savefig(path, dpi=180, bbox_inches="tight")
    plt.close(figure)


def choose_selected_k(
    selected_k_arg: str,
    cluster_counts: Sequence[int],
    evaluations: Sequence[ClusterEvaluation],
) -> int:
    if selected_k_arg == "auto":
        return max(evaluations, key=lambda item: item.silhouette_cosine).n_clusters
    try:
        selected = int(selected_k_arg)
    except ValueError as error:
        raise ValueError("--selected-k は auto または整数で指定してください。") from error
    if selected not in cluster_counts:
        raise ValueError(
            f"--selected-k={selected} は候補 {list(cluster_counts)} に含まれていません。"
        )
    return selected


def main() -> int:
    args = parse_args()
    input_path = args.input.expanduser().resolve()
    output_dir = args.output_dir.expanduser().resolve()
    if not input_path.is_file():
        print(f"入力ファイルがありません: {input_path}", file=sys.stderr)
        return 2
    output_dir.mkdir(parents=True, exist_ok=True)

    try:
        sheet_name, issues = read_issues(input_path, args.sheet, args.text_column)
        if len(issues) < 3:
            raise ValueError("空欄でない課題文が3件以上必要です。")
        cluster_counts = validate_cluster_counts(args.cluster_counts, len(issues))
        texts = [issue.text for issue in issues]
        embeddings = load_or_create_embeddings(
            texts=texts,
            model_name=args.model,
            output_dir=output_dir,
            batch_size=args.batch_size,
            device=args.device,
            force=args.force_embedding,
        )
        if embeddings.ndim != 2 or embeddings.shape[0] != len(issues):
            raise ValueError(
                f"埋め込みの形状が不正です: {embeddings.shape}, 課題数={len(issues)}"
            )

        all_labels: dict[int, np.ndarray] = {}
        evaluations: list[ClusterEvaluation] = []
        for count in cluster_counts:
            print(f"クラスタリング・評価中: k={count}")
            labels = fit_clusters(embeddings, count)
            all_labels[count] = labels
            evaluations.append(evaluate_clustering(embeddings, labels, count))

        selected_k = choose_selected_k(args.selected_k, cluster_counts, evaluations)
        print(f"結果Excelに採用するクラスタ数: k={selected_k}")
        coordinates = create_umap(
            embeddings, args.random_state, args.umap_neighbors
        )

        write_evaluation_csv(output_dir / "cluster_evaluation.csv", evaluations)
        write_assignments_csv(
            output_dir / "cluster_assignments_all.csv",
            issues,
            all_labels,
            coordinates,
        )
        write_cluster_summary(
            output_dir / "cluster_summary.xlsx",
            issues,
            embeddings,
            all_labels[selected_k],
            selected_k,
            args.representatives,
        )
        result_path = output_dir / f"{input_path.stem}_clustered_k{selected_k}.xlsx"
        append_clusters_to_excel(
            input_path,
            result_path,
            sheet_name,
            issues,
            all_labels[selected_k],
        )
        plot_comparison(
            output_dir / "umap_cluster_comparison.png",
            coordinates,
            all_labels,
            evaluations,
        )
        metadata = {
            "created_at_utc": datetime.now(timezone.utc).isoformat(),
            "input": str(input_path),
            "sheet": sheet_name,
            "text_column": args.text_column,
            "n_issues": len(issues),
            "model": args.model,
            "distance": "cosine",
            "linkage": "average",
            "cluster_counts": cluster_counts,
            "selected_k": selected_k,
            "random_state": args.random_state,
        }
        (output_dir / "run_metadata.json").write_text(
            json.dumps(metadata, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
    except (OSError, ValueError) as error:
        print(f"エラー: {error}", file=sys.stderr)
        return 2

    print(f"完了: {output_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
