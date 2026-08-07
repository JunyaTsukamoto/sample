import importlib.util
import sys
import tempfile
import unittest
from pathlib import Path
from unittest import mock

import numpy as np
from openpyxl import Workbook, load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
SPEC = importlib.util.spec_from_file_location(
    "cluster_issues", PROJECT_ROOT / "cluster_issues.py"
)
cluster_issues = importlib.util.module_from_spec(SPEC)
sys.modules[SPEC.name] = cluster_issues
assert SPEC.loader is not None
SPEC.loader.exec_module(cluster_issues)

NAMES_SPEC = importlib.util.spec_from_file_location(
    "apply_cluster_names", PROJECT_ROOT / "apply_cluster_names.py"
)
apply_cluster_names = importlib.util.module_from_spec(NAMES_SPEC)
sys.modules[NAMES_SPEC.name] = apply_cluster_names
assert NAMES_SPEC.loader is not None
NAMES_SPEC.loader.exec_module(apply_cluster_names)

LLM_SPEC = importlib.util.spec_from_file_location(
    "llm_classify", PROJECT_ROOT / "llm_classify.py"
)
llm_classify = importlib.util.module_from_spec(LLM_SPEC)
sys.modules[LLM_SPEC.name] = llm_classify
assert LLM_SPEC.loader is not None
LLM_SPEC.loader.exec_module(llm_classify)


class CoreTests(unittest.TestCase):
    def test_read_issues_skips_blank_cells_and_tracks_excel_rows(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "input.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "訓練の課題一覧"
            worksheet.append(["論文番号", "課題"])
            worksheet.append([1, "避難経路が不明確"])
            worksheet.append([2, None])
            worksheet.append([3, "  情報共有が遅れた  "])
            workbook.save(path)

            sheet, issues = cluster_issues.read_issues(path, None, "課題")

            self.assertEqual(sheet, "訓練の課題一覧")
            self.assertEqual([item.excel_row for item in issues], [2, 4])
            self.assertEqual(
                [item.text for item in issues],
                ["避難経路が不明確", "情報共有が遅れた"],
            )

    def test_clustering_and_representatives(self):
        embeddings = np.array(
            [
                [1.0, 0.0, 0.0],
                [0.9, 0.1, 0.0],
                [0.0, 1.0, 0.0],
                [0.1, 0.9, 0.0],
                [0.0, 0.0, 1.0],
                [0.0, 0.1, 0.9],
            ]
        )
        labels = cluster_issues.fit_clusters(embeddings, 3)
        evaluation = cluster_issues.evaluate_clustering(
            embeddings, labels, 3, min_cluster_size=2, max_cluster_share=0.5
        )
        representatives = cluster_issues.representative_indices(
            embeddings, labels, 1
        )

        self.assertEqual(len(set(labels)), 3)
        self.assertGreater(evaluation.silhouette_cosine, 0.8)
        self.assertTrue(evaluation.is_valid_candidate)
        self.assertEqual(set(representatives), set(labels))
        self.assertTrue(all(len(values) == 1 for values in representatives.values()))

    def test_append_clusters_preserves_other_columns(self):
        with tempfile.TemporaryDirectory() as directory:
            input_path = Path(directory) / "input.xlsx"
            output_path = Path(directory) / "output.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "data"
            worksheet.append(["ID", "課題"])
            worksheet.append([1, "a"])
            worksheet.append([2, "b"])
            workbook.save(input_path)
            issues = [
                cluster_issues.IssueRow(2, "a"),
                cluster_issues.IssueRow(3, "b"),
            ]

            cluster_issues.append_clusters_to_excel(
                input_path,
                output_path,
                "data",
                issues,
                np.array([1, 0]),
            )

            result = load_workbook(output_path, data_only=True)
            sheet = result["data"]
            self.assertEqual(sheet["A2"].value, 1)
            self.assertEqual(sheet["C1"].value, "cluster")
            self.assertEqual(sheet["D1"].value, "cluster_name")
            self.assertEqual(sheet["E1"].value, "is_outlier")
            self.assertEqual(sheet["F1"].value, "outlier_score")
            self.assertEqual(sheet["C2"].value, 1)
            self.assertEqual(sheet["C3"].value, 0)
            result.close()

    def test_candidate_constraints_and_selection(self):
        embeddings = np.array(
            [
                [1.0, 0.0],
                [0.9, 0.1],
                [0.8, 0.2],
                [0.7, 0.3],
                [0.0, 1.0],
            ]
        )
        labels = np.array([0, 0, 0, 0, 1])
        rejected = cluster_issues.evaluate_clustering(
            embeddings,
            labels,
            2,
            min_cluster_size=2,
            max_cluster_share=0.7,
        )

        self.assertFalse(rejected.is_valid_candidate)
        self.assertIn("smallest_cluster<2", rejected.rejection_reason)
        self.assertIn("largest_cluster_share>0.70", rejected.rejection_reason)
        self.assertIsNone(
            cluster_issues.choose_selected_k("auto", [2], [rejected])
        )
        self.assertEqual(
            cluster_issues.choose_selected_k("2", [2], [rejected]), 2
        )

    def test_detect_outliers_can_be_disabled(self):
        embeddings = np.eye(4)
        mask, scores = cluster_issues.detect_outliers(
            embeddings, "none", contamination=0.03, n_neighbors=20
        )

        self.assertFalse(mask.any())
        self.assertTrue(np.isnan(scores).all())

    def test_apply_cluster_names(self):
        with tempfile.TemporaryDirectory() as directory:
            directory_path = Path(directory)
            clustered_path = directory_path / "clustered.xlsx"
            summary_path = directory_path / "summary.xlsx"
            output_path = directory_path / "labeled.xlsx"

            clustered = Workbook()
            data_sheet = clustered.active
            data_sheet.append(["課題", "cluster", "cluster_name"])
            data_sheet.append(["a", 0, ""])
            data_sheet.append(["b", 1, ""])
            clustered.save(clustered_path)

            summary = Workbook()
            summary_sheet = summary.active
            summary_sheet.append(["cluster", "cluster_name"])
            summary_sheet.append([0, "情報共有"])
            summary_sheet.append([1, "避難誘導"])
            summary.save(summary_path)

            names = apply_cluster_names.read_cluster_names(summary_path)
            apply_cluster_names.apply_names(
                clustered_path, output_path, names, None
            )

            result = load_workbook(output_path, data_only=True)
            sheet = result[result.sheetnames[0]]
            self.assertEqual(sheet["C2"].value, "情報共有")
            self.assertEqual(sheet["C3"].value, "避難誘導")
            result.close()

    def test_llm_validation_adds_names_and_review_guardrail(self):
        taxonomy = {
            "categories": [
                {"code": "C09", "name": "医療・保健・福祉"},
                {"code": "C10", "name": "資機材・施設・システム"},
                {"code": "C13", "name": "その他・複合・判断困難"},
            ]
        }
        issues = [
            llm_classify.Issue(55, "使える施設をトリアージ的に選ぶ訓練が不足")
        ]
        raw = [{
            "excel_row": 55,
            "primary_code": "C09",
            "primary_subcategory": "救護・トリアージ",
            "secondary_codes": ["C13"],
            "issue_summary": "施設選定訓練の不足",
            "confidence": 0.85,
            "review_required": False,
            "rationale": "トリアージという語があるため",
        }]

        result = llm_classify.validate_batch(raw, issues, taxonomy)[0]

        self.assertEqual(result["primary_category"], "医療・保健・福祉")
        self.assertTrue(result["review_required"])
        self.assertTrue(result["rule_warnings"])

    def test_llm_low_confidence_requires_review(self):
        taxonomy = {"categories": [{"code": "C13", "name": "その他"}]}
        issues = [llm_classify.Issue(2, "情報が不足している")]
        raw = [{
            "excel_row": 2,
            "primary_code": "C13",
            "primary_subcategory": "情報不足",
            "secondary_codes": [],
            "issue_summary": "情報不足",
            "confidence": 0.70,
            "review_required": False,
            "rationale": "文脈が不足",
        }]

        self.assertTrue(llm_classify.validate_batch(raw, issues, taxonomy)[0]["review_required"])

    def test_llm_failed_large_batch_is_split(self):
        taxonomy = {"categories": [{"code": "C13", "name": "その他"}]}
        issues = [llm_classify.Issue(row, f"課題{row}") for row in range(2, 6)]

        def fake_call(endpoint, model, prompt, group, schema, timeout):
            selected = group[:1] if len(group) > 2 else group
            return [{
                "excel_row": issue.excel_row,
                "primary_code": "C13",
                "primary_subcategory": "情報不足",
                "secondary_codes": [],
                "confidence": 0.7,
                "review_required": True,
                "rationale": "情報不足",
            } for issue in selected]

        with mock.patch.object(llm_classify, "call_ollama", side_effect=fake_call), \
             mock.patch.object(llm_classify.time, "sleep"):
            results = llm_classify.classify_group(
                issues,
                endpoint="http://localhost",
                model="gemma4",
                prompt="prompt",
                schema={},
                taxonomy=taxonomy,
                timeout=1,
                retries=2,
            )

        self.assertEqual([item["excel_row"] for item in results], [2, 3, 4, 5])


if __name__ == "__main__":
    unittest.main()
