#!/usr/bin/env python3
"""課題文を主分類別に、原文フレーズ・対象―状態ペア・KWICで定量分析する。"""

from __future__ import annotations

import argparse
import csv
import json
import math
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Sequence

import ginza
import spacy
from openpyxl import load_workbook


GENERIC_TARGETS = {
    "こと", "もの", "ため", "ところ", "場合", "点", "上", "中", "等", "今回",
    "課題", "問題", "訓練", "対応", "実施", "検討", "必要", "状況", "結果",
}

# 課題内容を表す述語・状態語。原文表現は別列に保持し、ここでは集計用だけを正規化する。
STATE_CANONICAL = {
    "不足": "不足", "不足する": "不足", "欠如": "欠如", "欠ける": "欠如",
    "不十分": "不十分", "不明確": "不明確", "不明瞭": "不明確", "曖昧": "不明確",
    "不適切": "不適切", "不統一": "不統一", "不正確": "不正確", "不備": "不備",
    "困難": "困難", "難しい": "困難", "難い": "困難",
    "出来る": "できる", "できる": "できる", "行う": "実施", "実施する": "実施",
    "遅れる": "遅延", "遅延": "遅延", "時間がかかる": "時間がかかる", "かかる": "時間がかかる",
    "多い": "多い", "少ない": "少ない", "低い": "低い", "高い": "高い",
    "未確立": "未確立", "未実施": "未実施", "確立する": "確立",
    "確保": "確保が必要", "確保する": "確保が必要", "必要": "必要",
    "求める": "必要", "要する": "必要", "望ましい": "必要",
    "改善": "改善が必要", "改善する": "改善が必要", "見直し": "見直しが必要", "見直す": "見直しが必要",
    "徹底": "徹底が必要", "徹底する": "徹底が必要", "検討": "検討が必要", "検討する": "検討が必要",
    "懸念": "懸念", "負担": "負担", "混乱": "混乱", "混線": "混線",
    "ばらつき": "ばらつき", "異なる": "不一致", "依存": "依存", "偏る": "偏り",
    "限界": "限界", "不可能": "不可能", "できない": "できない", "ない": "ない",
}

NEGATION_LEMMAS = {"ない", "ぬ", "ず"}
NEGATIVE_CANONICAL = {
    "出来る": "できない", "できる": "できない", "分かる": "分からない",
    "確保": "確保できない", "かける": "かけられない", "高い": "高くない",
    "低い": "低くない", "明確": "不明確", "十分": "不十分", "なる": "ならない",
    "する": "しない", "即する": "即していない",
}
CONTENT_POS = {"NOUN", "PROPN", "PRON", "NUM"}
PREDICATE_POS = {"VERB", "ADJ", "NOUN"}


@dataclass(frozen=True)
class Issue:
    excel_row: int
    category: str
    text: str


@dataclass(frozen=True)
class Occurrence:
    excel_row: int
    category: str
    issue: str
    target_surface: str
    target_normalized: str
    state_surface: str
    state_normalized: str
    phrase_surface: str
    start_char: int
    end_char: int


def read_issues(
    path: Path,
    sheet_name: str | None,
    text_column: str,
    category_column: str,
) -> tuple[str, list[Issue]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name:
            worksheet = workbook[sheet_name]
        else:
            worksheet = next(
                sheet for sheet in workbook.worksheets
                if text_column in [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
                and category_column in [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
            )
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in worksheet[1]]
        missing = [name for name in (text_column, category_column) if name not in headers]
        if missing:
            raise ValueError(f"列がありません: {missing} / 利用可能な列: {headers}")
        text_index = headers.index(text_column)
        category_index = headers.index(category_column)
        issues: list[Issue] = []
        for excel_row, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            text = values[text_index]
            category = values[category_index]
            if text is None or not str(text).strip() or category is None or not str(category).strip():
                continue
            issues.append(Issue(excel_row, str(category).strip(), str(text).strip()))
        return worksheet.title, issues
    finally:
        workbook.close()


def clean_surface(text: str) -> str:
    return re.sub(r"^[\s、。・,:：;；（）()「」『』]+|[\s、。・,:：;；（）()「」『』]+$", "", text)


def compact_text(text: str) -> str:
    return re.sub(r"\s+", "", clean_surface(text))


def state_form(token) -> tuple[str, str] | None:
    """状態語の原文形と、集計用の最小限の正規形を返す。"""
    lemma = token.lemma_ or token.text
    auxiliaries = sorted(
        [child for child in token.children if child.dep_ in {"aux", "cop"}],
        key=lambda item: item.i,
    )
    negative = any((child.lemma_ or child.text) in NEGATION_LEMMAS for child in auxiliaries)
    if not negative and token.pos_ == "ADJ" and token.head is not token and token.head.pos_ in {"VERB", "ADJ"}:
        head_auxiliaries = [child for child in token.head.children if child.dep_ in {"aux", "cop"}]
        if any((child.lemma_ or child.text) in NEGATION_LEMMAS for child in head_auxiliaries):
            negative = True
            auxiliaries.extend(head_auxiliaries)
            auxiliaries.append(token.head)
    surface_tokens = [token, *auxiliaries]
    surface_tokens.sort(key=lambda item: item.i)
    surface = compact_text("".join(item.text_with_ws for item in surface_tokens))

    if negative:
        if lemma in NEGATIVE_CANONICAL:
            return surface, NEGATIVE_CANONICAL[lemma]
        if lemma == "明らか":
            return surface, "不明確"
        return surface, f"{lemma}ない"

    candidates = [lemma, token.text, surface]
    for candidate in candidates:
        if candidate in STATE_CANONICAL:
            return surface, STATE_CANONICAL[candidate]
    if token.pos_ in {"VERB", "ADJ"} and any(
        marker in surface for marker in ("必要", "困難", "不足", "できな", "難し", "遅れ")
    ):
        return surface, lemma
    return None


def noun_phrase(token) -> tuple[str, str, int, int] | None:
    """対象語を、連続する名詞・接頭辞を含む原文の名詞句として取得する。"""
    if token.pos_ not in CONTENT_POS:
        return None
    start = token.i
    end = token.i + 1
    cursor = token.i - 1
    while cursor >= token.sent.start:
        previous = token.doc[cursor]
        if previous.pos_ in CONTENT_POS or previous.dep_ in {"compound", "nmod"} or previous.pos_ == "PREFIX":
            start = cursor
            cursor -= 1
            continue
        break
    cursor = token.i + 1
    while cursor < token.sent.end:
        following = token.doc[cursor]
        if following.pos_ in CONTENT_POS and following.dep_ == "compound":
            end = cursor + 1
            cursor += 1
            continue
        break
    span = token.doc[start:end]
    surface = compact_text(span.text)
    normalized = "".join((item.lemma_ or item.text) for item in span if item.pos_ in CONTENT_POS or item.pos_ == "PREFIX")
    normalized = compact_text(normalized)
    if not surface or not normalized or normalized in GENERIC_TARGETS or len(normalized) < 2:
        return None
    return surface, normalized, span.start_char, span.end_char


def target_candidates(predicate) -> list:
    preferred = []
    fallback = []
    for child in predicate.children:
        if child.pos_ not in CONTENT_POS:
            continue
        if child.dep_ in {"nsubj", "obj", "obl", "nmod"}:
            preferred.append(child)
        else:
            fallback.append(child)
    if preferred:
        return sorted(preferred, key=lambda item: (abs(predicate.i - item.i), item.i))

    # 名詞述語では、係り先の名詞が対象になることがある。
    if predicate.head is not predicate and predicate.head.pos_ in CONTENT_POS:
        fallback.append(predicate.head)
    # 解析上直接つながらない短文に備え、同一文内で直前の名詞を最後の候補とする。
    preceding = [
        token for token in predicate.sent
        if token.i < predicate.i and token.pos_ in CONTENT_POS and (predicate.i - token.i) <= 8
    ]
    fallback.extend(reversed(preceding))
    seen = set()
    result = []
    for token in fallback:
        if token.i not in seen:
            result.append(token)
            seen.add(token.i)
    return result


def extract_occurrences(doc, issue: Issue) -> list[Occurrence]:
    occurrences: list[Occurrence] = []
    seen: set[tuple[str, str, int, int]] = set()
    for predicate in doc:
        if predicate.pos_ not in PREDICATE_POS:
            continue
        state = state_form(predicate)
        if state is None:
            continue
        for target in target_candidates(predicate):
            phrase = noun_phrase(target)
            if phrase is None:
                continue
            target_surface, target_normalized, target_start, target_end = phrase
            state_surface, state_normalized = state
            state_tokens = [predicate, *[child for child in predicate.children if child.dep_ in {"aux", "cop"}]]
            state_start = min(token.idx for token in state_tokens)
            state_end = max(token.idx + len(token.text) for token in state_tokens)
            start = min(target_start, state_start)
            end = max(target_end, state_end)
            phrase_surface = clean_surface(issue.text[start:end])
            if not phrase_surface or len(phrase_surface) > 80:
                continue
            key = (target_normalized, state_normalized, start, end)
            if key in seen:
                continue
            seen.add(key)
            occurrences.append(Occurrence(
                issue.excel_row,
                issue.category,
                issue.text,
                target_surface,
                target_normalized,
                state_surface,
                state_normalized,
                phrase_surface,
                start,
                end,
            ))
            break
    return occurrences


def document_sets(occurrences: Sequence[Occurrence], attribute: str) -> dict[tuple[str, str], set[int]]:
    result: dict[tuple[str, str], set[int]] = defaultdict(set)
    for occurrence in occurrences:
        result[(occurrence.category, getattr(occurrence, attribute))].add(occurrence.excel_row)
    return result


def pair_rows(
    occurrences: Sequence[Occurrence],
    category_sizes: dict[str, int],
    min_documents: int,
) -> list[dict]:
    pair_documents: dict[tuple[str, str, str], set[int]] = defaultdict(set)
    pair_occurrences: dict[tuple[str, str, str], list[Occurrence]] = defaultdict(list)
    target_documents = document_sets(occurrences, "target_normalized")
    state_documents = document_sets(occurrences, "state_normalized")
    for occurrence in occurrences:
        key = (occurrence.category, occurrence.target_normalized, occurrence.state_normalized)
        pair_documents[key].add(occurrence.excel_row)
        pair_occurrences[key].append(occurrence)

    rows = []
    for key, documents in pair_documents.items():
        category, target, state = key
        frequency = len(documents)
        if frequency < min_documents:
            continue
        n = category_sizes[category]
        target_frequency = len(target_documents[(category, target)])
        state_frequency = len(state_documents[(category, state)])
        pmi = math.log2((frequency * n) / (target_frequency * state_frequency))
        npmi_denominator = -math.log2(frequency / n) if frequency < n else 0
        npmi = pmi / npmi_denominator if npmi_denominator else 1.0
        dice = (2 * frequency) / (target_frequency + state_frequency)
        examples = pair_occurrences[key]
        surface_counter = Counter(item.phrase_surface for item in examples)
        representative = surface_counter.most_common(1)[0][0]
        rows.append({
            "AI主分類名": category,
            "対象（正規化）": target,
            "状態・述語（正規化）": state,
            "出現課題数": frequency,
            "主分類内課題数": n,
            "主分類内割合": frequency / n,
            "対象出現課題数": target_frequency,
            "状態出現課題数": state_frequency,
            "PMI": pmi,
            "nPMI": npmi,
            "Dice係数": dice,
            "代表原文フレーズ": representative,
            "Excel行": ";".join(str(row) for row in sorted(documents)),
        })
    rows.sort(key=lambda item: (item["AI主分類名"], -item["出現課題数"], -item["nPMI"], item["対象（正規化）"]))
    rank = Counter()
    for row in rows:
        rank[row["AI主分類名"]] += 1
        row["分類内順位"] = rank[row["AI主分類名"]]
    return rows


def phrase_rows(
    occurrences: Sequence[Occurrence],
    category_sizes: dict[str, int],
    pair_document_counts: dict[tuple[str, str, str], int],
) -> list[dict]:
    grouped: dict[tuple[str, str], list[Occurrence]] = defaultdict(list)
    for occurrence in occurrences:
        pair_key = (occurrence.category, occurrence.target_normalized, occurrence.state_normalized)
        if pair_key not in pair_document_counts:
            continue
        grouped[(occurrence.category, occurrence.phrase_surface)].append(occurrence)
    rows = []
    for (category, phrase), items in grouped.items():
        documents = {item.excel_row for item in items}
        pair_key = (category, items[0].target_normalized, items[0].state_normalized)
        rows.append({
            "AI主分類名": category,
            "原文フレーズ": phrase,
            "出現課題数": len(documents),
            "主分類内課題数": category_sizes[category],
            "主分類内割合": len(documents) / category_sizes[category],
            "対象―状態ペア出現課題数": pair_document_counts[pair_key],
            "対象（正規化）": items[0].target_normalized,
            "状態・述語（正規化）": items[0].state_normalized,
            "Excel行": ";".join(str(row) for row in sorted(documents)),
        })
    rows.sort(key=lambda item: (item["AI主分類名"], -item["出現課題数"], item["原文フレーズ"]))
    rank = Counter()
    for row in rows:
        rank[row["AI主分類名"]] += 1
        row["分類内順位"] = rank[row["AI主分類名"]]
    return rows


def kwic_rows(
    occurrences: Sequence[Occurrence],
    accepted_pairs: set[tuple[str, str, str]],
    context_chars: int,
) -> list[dict]:
    rows = []
    seen = set()
    for item in occurrences:
        key = (item.category, item.target_normalized, item.state_normalized)
        if key not in accepted_pairs:
            continue
        occurrence_key = (key, item.excel_row, item.start_char, item.end_char)
        if occurrence_key in seen:
            continue
        seen.add(occurrence_key)
        rows.append({
            "AI主分類名": item.category,
            "対象（正規化）": item.target_normalized,
            "状態・述語（正規化）": item.state_normalized,
            "Excel行": item.excel_row,
            "左文脈": item.issue[max(0, item.start_char - context_chars):item.start_char],
            "注目原文フレーズ": item.phrase_surface,
            "右文脈": item.issue[item.end_char:item.end_char + context_chars],
            "課題全文": item.issue,
        })
    rows.sort(key=lambda item: (item["AI主分類名"], item["対象（正規化）"], item["状態・述語（正規化）"], item["Excel行"]))
    return rows


def write_csv(path: Path, rows: Sequence[dict], fieldnames: Sequence[str]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("input_excel", type=Path)
    parser.add_argument("--output-dir", type=Path, default=Path("outputs_phrase_analysis"))
    parser.add_argument("--sheet")
    parser.add_argument("--text-column", default="課題")
    parser.add_argument("--category-column", default="AI主分類名")
    parser.add_argument("--model", default="ja_ginza")
    parser.add_argument("--min-documents", type=int, default=2)
    parser.add_argument("--context-chars", type=int, default=35)
    parser.add_argument("--batch-size", type=int, default=64)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    if args.min_documents < 1:
        raise ValueError("--min-documents は1以上にしてください。")
    sheet_name, issues = read_issues(
        args.input_excel, args.sheet, args.text_column, args.category_column
    )
    if not issues:
        raise ValueError("分析対象がありません。")
    print(f"対象シート: {sheet_name} / 課題: {len(issues)}件")
    nlp = spacy.load(args.model)
    occurrences = []
    for issue, doc in zip(
        issues,
        nlp.pipe((issue.text for issue in issues), batch_size=args.batch_size),
        strict=True,
    ):
        occurrences.extend(extract_occurrences(doc, issue))

    category_sizes = Counter(issue.category for issue in issues)
    pairs = pair_rows(occurrences, category_sizes, args.min_documents)
    accepted_pairs = {
        (row["AI主分類名"], row["対象（正規化）"], row["状態・述語（正規化）"])
        for row in pairs
    }
    pair_document_counts = {
        (row["AI主分類名"], row["対象（正規化）"], row["状態・述語（正規化）"]): row["出現課題数"]
        for row in pairs
    }
    phrases = phrase_rows(occurrences, category_sizes, pair_document_counts)
    kwic = kwic_rows(occurrences, accepted_pairs, args.context_chars)

    args.output_dir.mkdir(parents=True, exist_ok=True)
    write_csv(
        args.output_dir / "01_original_phrases.csv",
        phrases,
        ["AI主分類名", "分類内順位", "原文フレーズ", "出現課題数", "主分類内課題数", "主分類内割合", "対象―状態ペア出現課題数", "対象（正規化）", "状態・述語（正規化）", "Excel行"],
    )
    write_csv(
        args.output_dir / "02_target_state_pairs.csv",
        pairs,
        ["AI主分類名", "分類内順位", "対象（正規化）", "状態・述語（正規化）", "出現課題数", "主分類内課題数", "主分類内割合", "対象出現課題数", "状態出現課題数", "PMI", "nPMI", "Dice係数", "代表原文フレーズ", "Excel行"],
    )
    write_csv(
        args.output_dir / "03_kwic.csv",
        kwic,
        ["AI主分類名", "対象（正規化）", "状態・述語（正規化）", "Excel行", "左文脈", "注目原文フレーズ", "右文脈", "課題全文"],
    )
    write_csv(
        args.output_dir / "00_category_counts.csv",
        [{"AI主分類名": category, "課題数": count} for category, count in sorted(category_sizes.items())],
        ["AI主分類名", "課題数"],
    )
    metadata = {
        "input_excel": str(args.input_excel.resolve()),
        "sheet": sheet_name,
        "text_column": args.text_column,
        "category_column": args.category_column,
        "spacy_model": args.model,
        "issue_count": len(issues),
        "category_counts": dict(sorted(category_sizes.items())),
        "occurrence_count": len(occurrences),
        "phrase_row_count": len(phrases),
        "pair_row_count": len(pairs),
        "kwic_row_count": len(kwic),
        "min_documents": args.min_documents,
        "context_chars": args.context_chars,
        "normalization_policy": "原文フレーズは変更せず、対象の基本形と状態語だけを集計用に正規化",
    }
    (args.output_dir / "analysis_metadata.json").write_text(
        json.dumps(metadata, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(
        f"抽出: {len(occurrences)}箇所 / 原文フレーズ: {len(phrases)}行 / "
        f"対象―状態: {len(pairs)}行 / KWIC: {len(kwic)}行"
    )
    print(f"完了: {args.output_dir.resolve()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
