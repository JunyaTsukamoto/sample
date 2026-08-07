#!/usr/bin/env python3
"""Apply researcher-provided cluster names to a clustered Excel workbook."""

from __future__ import annotations

import argparse
from pathlib import Path
from shutil import copy2

from openpyxl import load_workbook


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="cluster_summary.xlsxのクラスタ名を結果Excelへ反映します。"
    )
    parser.add_argument("clustered_excel", type=Path)
    parser.add_argument("cluster_summary", type=Path)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--sheet", help="対象シート名。省略時は最初のシート")
    return parser.parse_args()


def header_map(worksheet) -> dict[str, int]:
    return {
        str(cell.value).strip(): cell.column
        for cell in worksheet[1]
        if cell.value is not None
    }


def read_cluster_names(summary_path: Path) -> dict[int, str]:
    workbook = load_workbook(summary_path, read_only=True, data_only=True)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        headers = header_map(worksheet)
        required = {"cluster", "cluster_name"}
        if not required.issubset(headers):
            raise ValueError(f"要約ファイルに必要な列がありません: {sorted(required)}")
        names: dict[int, str] = {}
        for row_number in range(2, worksheet.max_row + 1):
            cluster = worksheet.cell(row_number, headers["cluster"]).value
            name = worksheet.cell(row_number, headers["cluster_name"]).value
            if cluster is None or name is None or not str(name).strip():
                continue
            cluster_id = int(cluster)
            cleaned = str(name).strip()
            previous = names.get(cluster_id)
            if previous is not None and previous != cleaned:
                raise ValueError(
                    f"cluster={cluster_id} に異なる名前が複数あります: "
                    f"{previous!r}, {cleaned!r}"
                )
            names[cluster_id] = cleaned
        if not names:
            raise ValueError("cluster_nameが1件も入力されていません。")
        return names
    finally:
        workbook.close()


def apply_names(
    input_path: Path,
    output_path: Path,
    names: dict[int, str],
    sheet_name: str | None,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    copy2(input_path, output_path)
    workbook = load_workbook(output_path)
    try:
        worksheet = (
            workbook[workbook.sheetnames[0]]
            if sheet_name is None
            else workbook[sheet_name]
        )
        headers = header_map(worksheet)
        required = {"cluster", "cluster_name"}
        if not required.issubset(headers):
            raise ValueError(f"結果Excelに必要な列がありません: {sorted(required)}")
        for row_number in range(2, worksheet.max_row + 1):
            cluster = worksheet.cell(row_number, headers["cluster"]).value
            if cluster is not None and int(cluster) in names:
                worksheet.cell(
                    row_number,
                    headers["cluster_name"],
                    value=names[int(cluster)],
                )
        workbook.save(output_path)
    finally:
        workbook.close()


def main() -> int:
    args = parse_args()
    names = read_cluster_names(args.cluster_summary.expanduser().resolve())
    apply_names(
        args.clustered_excel.expanduser().resolve(),
        args.output.expanduser().resolve(),
        names,
        args.sheet,
    )
    print(f"反映完了: {args.output.expanduser().resolve()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
