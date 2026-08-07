#!/usr/bin/env python3
"""Gemma 4 + Ollamaで防災訓練の課題文を階層・複数ラベル分類する。"""

from __future__ import annotations

import argparse
import json
import sys
import time
import urllib.error
import urllib.request
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parent
DEFAULT_TAXONOMY = PROJECT_ROOT / "taxonomy.json"


@dataclass(frozen=True)
class Issue:
    excel_row: int
    text: str


def read_issues(path: Path, sheet_name: str | None, text_column: str) -> tuple[str, list[Issue]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in worksheet[1]]
        if text_column not in headers:
            raise ValueError(f"列 '{text_column}' がありません。列: {headers}")
        column = headers.index(text_column) + 1
        issues = []
        for row in range(2, worksheet.max_row + 1):
            value = worksheet.cell(row, column).value
            if value is not None and str(value).strip():
                issues.append(Issue(row, str(value).strip()))
        return worksheet.title, issues
    finally:
        workbook.close()


def batches(values: list[Issue], size: int) -> Iterable[list[Issue]]:
    for start in range(0, len(values), size):
        yield values[start : start + size]


def response_schema(valid_codes: list[str]) -> dict[str, Any]:
    item = {
        "type": "object",
        "properties": {
            "excel_row": {"type": "integer"},
            "primary_code": {"type": "string", "enum": valid_codes},
            "primary_subcategory": {"type": "string"},
            "secondary_codes": {
                "type": "array",
                "items": {"type": "string", "enum": valid_codes},
                "maxItems": 2,
            },
            "confidence": {"type": "number", "minimum": 0, "maximum": 1},
            "review_required": {"type": "boolean"},
            "rationale": {"type": "string"},
        },
        "required": [
            "excel_row", "primary_code", "primary_subcategory", "secondary_codes",
            "confidence", "review_required", "rationale"
        ],
        "additionalProperties": False,
    }
    return {
        "type": "object",
        "properties": {"classifications": {"type": "array", "items": item}},
        "required": ["classifications"],
        "additionalProperties": False,
    }


def build_system_prompt(taxonomy: dict[str, Any]) -> str:
    lines = [
        "あなたは防災訓練研究のコーディング担当です。課題文を暫定分類体系に沿って分類してください。",
        "目的は自動的に正解を確定することではなく、研究者の探索と再検討を支援することです。",
        "規則:",
        "1. 文面に明示された課題の中心をprimary_codeにする。原因や関連テーマはsecondary_codesに最大2件入れる。",
        "2. 書かれていない事実を推測しない。情報不足・複合・境界事例はreview_required=trueにする。",
        "3. confidenceが0.75未満の場合もreview_required=trueにする。",
        "4. 医療行為、トリアージ、傷病者対応は原則C09。単なる医療機関との調整はC07も検討する。",
        "   ただし『施設のトリアージ』『資源をトリアージ的に選ぶ』など人の傷病判定ではない比喩的用法はC09にしない。",
        "5. 情報の中身を集め整理する課題はC04、伝える経路・報告・通信はC05。",
        "6. 訓練当日の進行はC02、訓練の目的・シナリオ設計はC01。",
        "7. 評価・検証を行う方法や基準そのものが中心ならC03。評価用データの収集・入力が中心ならC04/C05/C10を検討する。",
        "8. confidenceは境界事例0.55〜0.74、概ね明確0.75〜0.89、ほぼ一義的0.90〜0.97とし、1.0は使わない。",
        "9. rationaleは分類判断の決め手だけを日本語20字以内で記す。",
        "10. 入力したexcel_rowを変更せず、すべての入力に1件ずつ結果を返す。",
        "分類体系:",
        json.dumps(taxonomy["categories"], ensure_ascii=False),
    ]
    return "\n".join(lines)


def call_ollama(
    endpoint: str,
    model: str,
    system_prompt: str,
    issues: list[Issue],
    schema: dict[str, Any],
    timeout: int,
) -> list[dict[str, Any]]:
    payload = {
        "model": model,
        "stream": False,
        "format": schema,
        "messages": [
            {"role": "system", "content": system_prompt},
            {
                "role": "user",
                "content": json.dumps(
                    [{"excel_row": i.excel_row, "issue": i.text} for i in issues],
                    ensure_ascii=False,
                ),
            },
        ],
        "options": {"temperature": 0, "seed": 42, "num_ctx": 16384},
        "keep_alive": "20m",
    }
    request = urllib.request.Request(
        endpoint.rstrip("/") + "/api/chat",
        data=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    with urllib.request.urlopen(request, timeout=timeout) as response:
        envelope = json.loads(response.read().decode("utf-8"))
    content = envelope["message"]["content"]
    return json.loads(content)["classifications"]


def validate_batch(
    raw: list[dict[str, Any]], issues: list[Issue], taxonomy: dict[str, Any]
) -> list[dict[str, Any]]:
    expected = {issue.excel_row: issue for issue in issues}
    categories = {item["code"]: item for item in taxonomy["categories"]}
    by_row: dict[int, dict[str, Any]] = {}
    for result in raw:
        row = int(result["excel_row"])
        if row not in expected or row in by_row:
            raise ValueError(f"予期しない、または重複したexcel_row: {row}")
        primary = result["primary_code"]
        secondary = [code for code in result.get("secondary_codes", []) if code != primary]
        if primary not in categories or any(code not in categories for code in secondary):
            raise ValueError(f"無効な分類コード: row={row}")
        confidence = float(result["confidence"])
        text = expected[row].text
        warnings: list[str] = []
        metaphorical_triage = (
            "トリアージ" in text
            and any(word in text for word in ("施設", "資源", "物資"))
            and not any(word in text for word in ("タグ", "傷病", "患者", "救護"))
        )
        if metaphorical_triage and primary == "C09":
            warnings.append("比喩的なトリアージを医療分類にした可能性")
        if any(word in text for word in ("傷病者", "トリアージタグ", "医療救護")) and primary != "C09":
            warnings.append("明示的な医療語があるが主分類がC09ではない")
        if "無線" in text and primary != "C05" and "C05" not in secondary:
            warnings.append("無線の課題だがC05が主・副分類にない")
        review_reasons = list(warnings)
        if confidence < 0.90:
            review_reasons.append("AI信頼度0.90未満")
        if primary == "C13":
            review_reasons.append("その他・複合・判断困難（C13）")
        if len(secondary) >= 2:
            review_reasons.append("副分類が2件ある複合課題")
        review_required = bool(result["review_required"]) or bool(review_reasons)
        review_priority = (
            "高" if warnings or confidence < 0.80 else "中" if review_required else "通常"
        )
        by_row[row] = {
            "excel_row": row,
            "issue": expected[row].text,
            "primary_code": primary,
            "primary_category": categories[primary]["name"],
            "primary_subcategory": str(result["primary_subcategory"]).strip(),
            "secondary_codes": secondary[:2],
            "secondary_categories": [categories[code]["name"] for code in secondary[:2]],
            "confidence": confidence,
            "review_required": review_required,
            "review_priority": review_priority,
            "review_reasons": review_reasons,
            "rationale": str(result["rationale"]).strip(),
            "rule_warnings": warnings,
        }
    missing = sorted(set(expected) - set(by_row))
    if missing:
        raise ValueError(f"結果がないexcel_row: {missing}")
    return [by_row[issue.excel_row] for issue in issues]


def load_completed(path: Path) -> dict[int, dict[str, Any]]:
    if not path.exists():
        return {}
    completed = {}
    with path.open(encoding="utf-8") as handle:
        for line in handle:
            if line.strip():
                item = json.loads(line)
                completed[int(item["excel_row"])] = item
    return completed


def classify_group(
    issues: list[Issue],
    *,
    endpoint: str,
    model: str,
    prompt: str,
    schema: dict[str, Any],
    taxonomy: dict[str, Any],
    timeout: int,
    retries: int,
) -> list[dict[str, Any]]:
    """バッチを分類し、失敗が続けば半分に分割して完全な結果だけを返す。"""
    last_error: Exception | None = None
    for attempt in range(1, retries + 1):
        try:
            raw = call_ollama(endpoint, model, prompt, issues, schema, timeout)
            return validate_batch(raw, issues, taxonomy)
        except (urllib.error.URLError, TimeoutError, KeyError, ValueError, json.JSONDecodeError) as error:
            last_error = error
            print(f"再試行 {attempt}/{retries}: {error}", file=sys.stderr, flush=True)
            time.sleep(min(2 ** attempt, 8))
    if len(issues) == 1:
        raise RuntimeError(f"分類に失敗したExcel行: {issues[0].excel_row}") from last_error
    midpoint = len(issues) // 2
    print(
        f"バッチを{midpoint}件と{len(issues)-midpoint}件に分割して再試行",
        file=sys.stderr,
        flush=True,
    )
    common = {
        "endpoint": endpoint,
        "model": model,
        "prompt": prompt,
        "schema": schema,
        "taxonomy": taxonomy,
        "timeout": timeout,
        "retries": retries,
    }
    return classify_group(issues[:midpoint], **common) + classify_group(issues[midpoint:], **common)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("input_excel", type=Path)
    parser.add_argument("--output", type=Path, default=Path("outputs_llm/gemma4_classifications.jsonl"))
    parser.add_argument("--taxonomy", type=Path, default=DEFAULT_TAXONOMY)
    parser.add_argument("--model", default="gemma4:latest")
    parser.add_argument("--endpoint", default="http://127.0.0.1:11434")
    parser.add_argument("--sheet")
    parser.add_argument("--text-column", default="課題")
    parser.add_argument("--batch-size", type=int, default=6)
    parser.add_argument("--limit", type=int)
    parser.add_argument("--rows", type=int, nargs="+")
    parser.add_argument("--timeout", type=int, default=300)
    parser.add_argument("--retries", type=int, default=3)
    parser.add_argument("--overwrite", action="store_true")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    taxonomy = json.loads(args.taxonomy.read_text(encoding="utf-8"))
    sheet, issues = read_issues(args.input_excel, args.sheet, args.text_column)
    if args.rows:
        requested = set(args.rows)
        issues = [issue for issue in issues if issue.excel_row in requested]
    if args.limit is not None:
        issues = issues[: args.limit]
    if not issues:
        raise ValueError("分類対象の課題がありません。")

    args.output.parent.mkdir(parents=True, exist_ok=True)
    if args.overwrite and args.output.exists():
        args.output.unlink()
    completed = load_completed(args.output)
    pending = [issue for issue in issues if issue.excel_row not in completed]
    codes = [item["code"] for item in taxonomy["categories"]]
    schema = response_schema(codes)
    prompt = build_system_prompt(taxonomy)

    print(f"対象シート: {sheet} / 対象: {len(issues)}件 / 再開済み: {len(issues)-len(pending)}件")
    with args.output.open("a", encoding="utf-8") as handle:
        done = len(issues) - len(pending)
        for group in batches(pending, args.batch_size):
            results = classify_group(
                group,
                endpoint=args.endpoint,
                model=args.model,
                prompt=prompt,
                schema=schema,
                taxonomy=taxonomy,
                timeout=args.timeout,
                retries=args.retries,
            )
            for result in results:
                handle.write(json.dumps(result, ensure_ascii=False) + "\n")
            handle.flush()
            done += len(group)
            print(f"分類済み: {done}/{len(issues)}", flush=True)
    print(f"完了: {args.output.resolve()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
